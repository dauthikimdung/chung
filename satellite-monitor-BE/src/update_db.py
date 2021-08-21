from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.common import exceptions
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import openpyxl
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import csv
import urllib.request
from pymongo import MongoClient
url = 'http://celestrak.com/NORAD/elements/active.txt' # Satellite Database online
dbUrl = "mongodb+srv://satelliteV10:7FgUH3CrGH21@satellite0.fvo32.mongodb.net/satellite?retryWrites=true&w=majority"
dbName = "satellite"
collName = "full"
client = MongoClient(dbUrl, ssl=True, ssl_cert_reqs='CERT_NONE')
db = client[dbName]
coll = db[collName]
timeout = 40
def findByXpath(driver, xpath):
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xpath))).text
# Chuẩn hóa chuỗi ký tự (bỏ các ký tự thừa: \n, \t, "  ")
def nomalizeString(str):
    return " ".join(str.split())
def mongoImport(csvPath):
    """ Imports a csv file at path csvPath to a mongo collection
    returns: count of the documents in the new collection
    """
    global coll # collection
    csvFile = open(csvPath)
    reader = csv.DictReader(csvFile)
    count = 0 
    for each in reader:
        row={}
        for field in header:
            row[field]= " ".join(each[field].split())
        print(row)
        coll.insert(row)
        count+=1
    return count
options = webdriver.ChromeOptions()
options.add_argument('--headless')
options.add_argument('--no-sandbox')
driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options) #desired_capabilities=capabilities
# Đọc Norad Number offline
# wb = openpyxl.load_workbook('..\\DB_of_STL.xlsx')
# sheet=wb.active
# max_row=sheet.max_row
# listID=[]
# for i in range(2,max_row+1):
#     value_id = sheet.cell(i, 2).value
#     listID.append(value_id) # Creat list of id of excel file
# f = open('data.txt'); list_content=f.readlines()
# Tạo file csv
csvFile = open('updated_data.csv', 'w', newline='', encoding='utf-8')
writer = csv.writer(csvFile) # công cụ ghi file csv
emptyRowCSV = [""] * 17  # Dòng dữ liệu trống dùng làm mẫu
# Dòng tên trường
header= ["Official Name","NORAD Number","Nation","Operator","Users","Application",
"Detailed Purpose","Orbit","Class of Orbit","Type of Orbit","Period (minutes)","Mass (kg)",
"COSPAR Number","Date of Launch","Expected Lifetime (yrs)","Equipment","Describe"]
writer.writerow(header)

local_filename, headers = urllib.request.urlretrieve(url,filename="..\\data.txt",)
f = open(local_filename, encoding="utf-8")
list_content=f.readlines()
for i in range(2, len(list_content),3):
    id_str=list_content[i][2:7]
    id_int=int(id_str)
    row = emptyRowCSV.copy() # Tạo bản ghi dữ liệu mới (dòng mới)
    if id_int==45123 or id_int==45125:
        continue
    if coll.find_one({'NORAD Number': id_int}) == None: # id_int not in listID
        print('New statellite:', id_int)
        row[1] = int(id_int)
        row[0] = nomalizeString(list_content[i - 2])
        if list_content[i-2][0:2] == '20':  # in case of name_satellite starting by 2021
            try:
                print("n2yo")
                driver.get('https://www.n2yo.com/')
                element_search = driver.find_element_by_name('q')
                element_search.send_keys(list_content[i-2])
                element_search.submit()
            except exceptions.StaleElementReferenceException:
                try:
                    infor_stl = findByXpath(driver,
                        "/html/body/table[@id='tabsatellite']/tbody/tr/td[2]/div[@id='satinfo']").splitlines()
                except (TimeoutException, NoSuchElementException):
                    print('TimeoutException')
                    continue
                name_stl = infor_stl[0][::]
                nat_stl = infor_stl[13][8:]
                orbit_stl = infor_stl[6][9:] + 'x' + infor_stl[7][8:] + 'x' + infor_stl[8][13:]
                cospar_number_stl = infor_stl[5][12:]
                launch_date_stl = infor_stl[12][13:]
                period_stl = infor_stl[9][8:]

                row[2] = nomalizeString(nat_stl)
                row[7] = nomalizeString(orbit_stl)
                row[0] = nomalizeString(name_stl)
                row[12] = nomalizeString(cospar_number_stl)
                row[10] = nomalizeString(period_stl)
                row[13] = nomalizeString(launch_date_stl)
        else:  # in case of normal name_satellite
            try:
                print("google")
                driver.get('https://www.google.com.vn/')
                element = driver.find_element_by_name('q')
                print('site:space.skyrocket.de ' + f'{list_content[i-2]}')
                element.send_keys('site:space.skyrocket.de ' + f'{list_content[i-2]}')  # search in space.skyrocket.de
                element.submit()
            except exceptions.StaleElementReferenceException:
                try:
                    result = driver.find_element_by_tag_name('h3')  # click on first result
                    result.click()
                    infor = findByXpath(driver,"/html/body/div[@class='page_bg']/div[@class='container']/div/div[@id='satdescription']/p[1]")
                    nat_stl = findByXpath(driver,
                        "/html/body/div[@class='page_bg']/div[@class='container']/div/table[@id='satdata']/tbody/tr[1]/td[@id='sdnat']")
                    typ_stl = findByXpath(driver,
                        "/html/body/div[@class='page_bg']/div[@class='container']/div/table[@id='satdata']/tbody/tr[2]/td[@id='sdtyp']")
                    ope_stl = findByXpath(driver,
                        "/html/body/div[@class='page_bg']/div[@class='container']/div/table[@id='satdata']/tbody/tr[3]/td[@id='sdope']")
                    equ_stl = findByXpath(driver,
                        "/html/body/div[@class='page_bg']/div[@class='container']/div/table[@id='satdata']/tbody/tr[5]/td[@id='sdequ']")
                    lif_stl = findByXpath(driver,
                        "/html/body/div[@class='page_bg']/div[@class='container']/div/table[@id='satdata']/tbody/tr[9]/td[@id='sdlif']")
                    mass_stl = findByXpath(driver,
                        "/html/body/div[@class='page_bg']/div[@class='container']/div/table[@id='satdata']/tbody/tr[10]/td[@id='sdmas']")
                    orbit_stl = findByXpath(driver,
                        "/html/body/div[@class='page_bg']/div[@class='container']/div/table[@id='satdata']/tbody/tr[11]/td[@id='sdorb']")
                    row[2] = nomalizeString(nat_stl)
                    row[3] = nomalizeString(ope_stl)
                    row[5] = nomalizeString(typ_stl)
                    row[7] = nomalizeString(orbit_stl)
                    row[11] = nomalizeString(mass_stl)
                    row[14] = nomalizeString(lif_stl)
                    row[15] = nomalizeString(equ_stl)
                    row[16] = nomalizeString(infor)
                except (TimeoutException, NoSuchElementException):
                    continue
        writer.writerow(row)
driver.close()
print(mongoImport(csvPath='updated_data.csv'))